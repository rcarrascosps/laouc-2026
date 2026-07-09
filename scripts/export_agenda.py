"""Export the public, confirmed-only LAOUC Tour agenda to JSON for the MCP server."""
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

GREEN_FILL = 'FFC6EFCE'

TOUR_DIR = Path(r"C:\rolando\SPS\2026\LAOUC\tour")
AGENDA_XLSX = TOUR_DIR / "agenda_laouc_tour_2026.xlsx"
SESIONES_XLSX = TOUR_DIR / "sesiones.xlsx"
ACCEPTED_CSV = TOUR_DIR / "accepted_sessions.csv"
OUTPUT_JSON = Path(__file__).resolve().parent.parent / "data" / "agenda-publica.json"

CITIES = ['Mexico', 'Guatemala', 'Costa Rica', 'Panama', 'Chile',
          'Brazil', 'Uruguay', 'Argentina', 'Paraguay']

# Official tour stop date per city (organizer-confirmed, 2026-06-23). Each city's
# sessions all happen on this single calendar day, in that city's local time.
CITY_DATES = {
    'Mexico': '2026-08-14',
    'Guatemala': '2026-08-17',
    'Costa Rica': '2026-08-19',
    'Panama': '2026-08-21',
    'Chile': '2026-08-24',
    'Paraguay': '2026-08-26',
    'Brazil': '2026-08-29',
    'Uruguay': '2026-08-31',
    'Argentina': '2026-09-02',
}


def is_confirmed(fill_rgb):
    return fill_rgb == GREEN_FILL


def parse_session_cell(text):
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if not lines:
        return {'title': '', 'speaker_name': ''}
    title = lines[0]
    speaker_line = next((l for l in lines[1:] if not l.startswith('(orig')), '')
    speaker_name = re.sub(r'\s*\(KEYNOTE\)\s*', '', speaker_line)
    speaker_name = re.sub(r'\s*\[.*?\]\s*', '', speaker_name).strip()
    return {'title': title, 'speaker_name': speaker_name}


def clean_str(value):
    if value is None:
        return ''
    if isinstance(value, float) and value != value:  # NaN
        return ''
    return str(value)


def extract_ace(tagline):
    if not tagline:
        return ''
    m = re.search(r'Oracle ACE\s*(Director|Pro|Associate|Apprentice)?', tagline, re.IGNORECASE)
    if not m:
        return ''
    tier = (m.group(1) or '').strip()
    return f'Oracle ACE {tier}'.strip() if tier else 'Oracle ACE'


def extract_company(tagline):
    if not tagline:
        return ''
    tl_clean = re.sub(
        r',?\s*Oracle ACE\s*(?:Director|Pro|Associate|Apprentice)?\s*,?', '', tagline, flags=re.IGNORECASE
    )
    tl_clean = tl_clean.strip().strip(',').strip(' |-').strip()
    if not tl_clean:
        return ''
    if ' - ' in tl_clean:
        return tl_clean.split(' - ')[0].strip()
    m = re.search(r'\bat\s+(.+?)(?:\s*[|,]|$)', tl_clean, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    if ',' in tl_clean:
        first = tl_clean.split(',')[0].strip()
        if len(first.split()) <= 4 and not first[0].islower():
            return first
    if len(tl_clean.split()) <= 3:
        return tl_clean
    return ''


def normalize_name(name):
    return re.sub(r'\s+', ' ', name).strip().casefold()


def build_speaker_lookup(accepted_rows, orig_rows):
    orig_by_id = {r['SessionId']: r for r in orig_rows}
    lookup = {}
    for r in accepted_rows:
        full_name = f"{r['FirstName']} {r['LastName']}".strip()
        key = normalize_name(full_name)
        if key in lookup:
            continue
        orig = orig_by_id.get(r['SessionId'], {})
        tagline = clean_str(orig.get('TagLine'))
        lookup[key] = {
            'company': extract_company(tagline),
            'bio': clean_str(orig.get('Bio')),
            'oracle_ace': extract_ace(tagline),
        }
    return lookup


# Distinguishes the keynote row (merged B3:E3, min_col=2) from the lunch-break
# row (merged A8:E8, min_col=1) — only a merge starting at column B is a keynote.
def is_keynote_row(ws, row_idx):
    for rng in ws.merged_cells.ranges:
        if rng.min_row == row_idx and rng.min_col == 2 and rng.max_col == 5:
            return True
    return False


def find_header_row(ws):
    for row in range(1, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value and re.match(r'(?i)^hor[aá]rio$', str(value).strip()):
            return row
    raise ValueError(f"No 'Horario' header row found in sheet '{ws.title}'")


# Custom-format cells list the speaker on the first line and the session title on
# the second (opposite of the standard template). A keynote is marked inline with
# a "Keynote:" prefix on the title line instead of a merged row + "(KEYNOTE)" tag.
# Single-line cells (Registro, Apertura, Coffee Break, Lunch Break, ...) are venue
# logistics, not sessions, and are skipped.
def parse_custom_format_cell(text):
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    if len(lines) < 2:
        return None
    speaker_name = lines[0]
    title_line = lines[1]
    keynote_match = re.match(r'(?i)^keynote:\s*', title_line)
    title = re.sub(r'(?i)^keynote:\s*', '', title_line).strip()
    return {'title': title, 'speaker_name': speaker_name, 'is_keynote': bool(keynote_match)}


def extract_custom_format_sessions(ws):
    city = ws.title
    header_row = find_header_row(ws)
    track_names = {col: ws.cell(row=header_row, column=col).value for col in range(2, 6)}
    entries = []
    for row in range(header_row + 1, ws.max_row + 1):
        time_slot = ws.cell(row=row, column=1).value
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                continue
            parsed = parse_custom_format_cell(cell.value)
            if parsed is None or not parsed['title']:
                continue
            keynote = parsed['is_keynote']
            entries.append({
                'city': city,
                'time_slot': None if keynote else time_slot,
                'track': None if keynote else track_names[col],
                'is_keynote': keynote,
                'title': parsed['title'],
                'speaker_name': parsed['speaker_name'],
                'fill_rgb': GREEN_FILL,
            })
    return entries


# Chile's cells pack "Title<many spaces>Speaker" onto a single line instead of two
# lines, and the fill colors are decorative per-track branding (a session moved from
# its home track keeps that track's color, like Mexico's "(orig: ...)" convention) —
# not a confirmation signal. Every parsed session is treated as confirmed.
CHILE_LOGISTICS = {'coffee break', 'closing'}


def parse_chile_regular_cell(text):
    text = text.strip()
    if text.lower() in CHILE_LOGISTICS:
        return None
    parts = [p.strip() for p in re.split(r'\s{2,}', text) if p.strip()]
    if len(parts) < 2:
        return None
    return {'title': ' '.join(parts[:-1]).strip(), 'speaker_name': parts[-1]}


def parse_chile_keynote_cell(text):
    text = re.sub(r'^[▶\s]*KEYNOTE\s*[—-]\s*', '', text, flags=re.IGNORECASE).strip()
    if ' - ' in text:
        title, speaker = text.rsplit(' - ', 1)
    else:
        title, speaker = text, ''
    return {'title': title.strip(), 'speaker_name': speaker.strip()}


def extract_chile_sessions(ws):
    city = ws.title
    header_row = find_header_row(ws)
    track_names = {col: ws.cell(row=header_row, column=col).value for col in range(2, 6)}
    entries = []
    for row in range(header_row + 1, ws.max_row + 1):
        time_slot = ws.cell(row=row, column=1).value
        keynote = is_keynote_row(ws, row)
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                continue
            text = str(cell.value)
            if text.strip().lower() in CHILE_LOGISTICS:
                continue
            parsed = parse_chile_keynote_cell(text) if keynote else parse_chile_regular_cell(text)
            if parsed is None or not parsed['title']:
                continue
            entries.append({
                'city': city,
                'time_slot': None if keynote else time_slot,
                'track': None if keynote else track_names[col],
                'is_keynote': keynote,
                'title': parsed['title'],
                'speaker_name': parsed['speaker_name'],
                'fill_rgb': GREEN_FILL,
            })
    return entries


# Brazil's cells pack Speaker(s) / Role-Company / Title / Language / optional Room
# on separate lines, with no confirmation-color convention either (everything on
# the sheet is final). Co-presented sessions (e.g. "Mike Dietrich & Harsh Gupta")
# get an extra role line per presenter, so the title can't be found by a fixed
# line index — anchor on the trailing language marker instead, which is always the
# line right before the title (or two lines before it, if a room line follows).
BRAZIL_LANGUAGE_MARKERS = {
    'português', 'portugues', 'portugês', 'inglês', 'ingles', 'english', 'espanhol', 'español', 'spanish',
}


def parse_brazil_cell(text):
    lines = [l.strip() for l in text.split('\n') if l.strip()]
    is_keynote = bool(lines) and lines[0].strip().lower().replace(' ', '') == 'keynote'
    if is_keynote:
        lines = lines[1:]
    if len(lines) < 4:
        return None
    speaker_name = lines[0]
    if lines[-1].strip().lower() in BRAZIL_LANGUAGE_MARKERS:
        title = lines[-2]
    elif len(lines) >= 5 and lines[-2].strip().lower() in BRAZIL_LANGUAGE_MARKERS:
        title = lines[-3]
    else:
        title = lines[-2]
    return {'title': title.strip(), 'speaker_name': speaker_name.strip(), 'is_keynote': is_keynote}


def extract_brazil_sessions(ws):
    city = ws.title
    header_row = find_header_row(ws)
    track_names = {col: ws.cell(row=header_row, column=col).value for col in range(2, 6)}
    entries = []
    for row in range(header_row + 1, ws.max_row + 1):
        time_slot = ws.cell(row=row, column=1).value
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                continue
            parsed = parse_brazil_cell(str(cell.value))
            if parsed is None or not parsed['title']:
                continue
            keynote = parsed['is_keynote']
            entries.append({
                'city': city,
                'time_slot': None if keynote else time_slot,
                'track': None if keynote else track_names[col],
                'is_keynote': keynote,
                'title': parsed['title'],
                'speaker_name': parsed['speaker_name'],
                'fill_rgb': GREEN_FILL,
            })
    return entries


# Cities whose sheet is a locally-published final agenda in its own custom layout
# (not the standard 4-track "Title\nSpeaker" + green/yellow confirmation template).
# Everything on these sheets is already final, so every parsed session is confirmed.
CITY_EXTRACTORS = {
    'Uruguay': extract_custom_format_sessions,
    'Chile': extract_chile_sessions,
    'Brazil': extract_brazil_sessions,
}


def extract_city_sessions(ws):
    city = ws.title
    header_row = find_header_row(ws)
    track_names = {col: ws.cell(row=header_row, column=col).value for col in range(2, 6)}
    entries = []
    for row in range(header_row + 1, ws.max_row + 1):
        time_slot = ws.cell(row=row, column=1).value
        keynote = is_keynote_row(ws, row)
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            if not cell.value:
                continue
            parsed = parse_session_cell(cell.value)
            entries.append({
                'city': city,
                'time_slot': None if keynote else time_slot,
                'track': None if keynote else track_names[col],
                'is_keynote': keynote,
                'title': parsed['title'],
                'speaker_name': parsed['speaker_name'],
                'fill_rgb': cell.fill.fgColor.rgb,
            })
    return entries


def build_public_sessions(raw_entries, speaker_lookup):
    public_sessions = []
    for entry in raw_entries:
        if not is_confirmed(entry['fill_rgb']):
            continue
        info = speaker_lookup.get(normalize_name(entry['speaker_name']), {})
        public_sessions.append({
            'city': entry['city'],
            'time_slot': entry['time_slot'],
            'track': entry['track'],
            'is_keynote': entry['is_keynote'],
            'title': entry['title'],
            'speaker_name': entry['speaker_name'],
            'speaker_company': info.get('company', ''),
            'speaker_bio': info.get('bio', ''),
            'oracle_ace': info.get('oracle_ace', '') or None,
        })
    return public_sessions


def main():
    accepted = pd.read_csv(ACCEPTED_CSV, dtype={'SessionId': str}).to_dict('records')
    orig_df = pd.read_excel(
        SESIONES_XLSX, sheet_name='Sessions and speakers - Origina', dtype={'Session Id': str}
    )
    orig_df = orig_df.rename(columns={'Session Id': 'SessionId'})
    orig_rows = orig_df.to_dict('records')

    speaker_lookup = build_speaker_lookup(accepted, orig_rows)

    wb = load_workbook(AGENDA_XLSX)
    all_public_sessions = []
    for city in CITIES:
        ws = wb[city]
        extractor = CITY_EXTRACTORS.get(city, extract_city_sessions)
        raw_entries = extractor(ws)
        city_sessions = build_public_sessions(raw_entries, speaker_lookup)
        for session in city_sessions:
            session['date'] = CITY_DATES[city]
        all_public_sessions.extend(city_sessions)

    output = {
        'generated_at': datetime.now(timezone.utc).isoformat(),
        'cities': CITIES,
        'city_dates': CITY_DATES,
        'sessions': all_public_sessions,
    }
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding='utf-8')
    print(f'Wrote {len(all_public_sessions)} sessions -> {OUTPUT_JSON}')


if __name__ == '__main__':
    sys.exit(main())
