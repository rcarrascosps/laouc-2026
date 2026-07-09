from export_agenda import is_confirmed, parse_session_cell, GREEN_FILL


def test_is_confirmed_green():
    assert is_confirmed(GREEN_FILL) is True


def test_is_confirmed_yellow_red_white():
    assert is_confirmed('FFFFF2CC') is False
    assert is_confirmed('FFFFC7CE') is False
    assert is_confirmed('FFF5F5F5') is False


def test_parse_session_cell_keynote():
    text = 'Usando Inteligencia Artificial CON Oracle AI Database\nEugenio Galiano  (KEYNOTE)'
    assert parse_session_cell(text) == {
        'title': 'Usando Inteligencia Artificial CON Oracle AI Database',
        'speaker_name': 'Eugenio Galiano',
    }


def test_parse_session_cell_with_orig_tag():
    text = "How to use a relational database for your JSON documents\nPatrick Barel\n(orig: Base de Datos)"
    assert parse_session_cell(text) == {
        'title': 'How to use a relational database for your JSON documents',
        'speaker_name': 'Patrick Barel',
    }


def test_parse_session_cell_with_ace_bracket():
    text = (
        'Real-Time AI: Consolidating Vector Data with OGG 26ai\n'
        'Gilson Martins  [Oracle ACE Director]\n(orig: Base de Datos)'
    )
    assert parse_session_cell(text) == {
        'title': 'Real-Time AI: Consolidating Vector Data with OGG 26ai',
        'speaker_name': 'Gilson Martins',
    }


def test_parse_session_cell_no_speaker_line():
    text = '▶  KEYNOTE  —  PENDIENTE DE CONFIRMAR'
    assert parse_session_cell(text) == {
        'title': 'PENDIENTE DE CONFIRMAR',
        'speaker_name': '',
    }


def test_parse_session_cell_strips_keynote_marker_prefix():
    text = '▶  KEYNOTE  —  The High Availability Mindset\nFrancisco Muñoz Alvarez'
    assert parse_session_cell(text) == {
        'title': 'The High Availability Mindset',
        'speaker_name': 'Francisco Muñoz Alvarez',
    }


def test_parse_session_cell_empty_text_returns_empty_fields():
    assert parse_session_cell('') == {'title': '', 'speaker_name': ''}
    assert parse_session_cell('   \n   ') == {'title': '', 'speaker_name': ''}


from export_agenda import (
    clean_str,
    extract_ace,
    extract_company,
    normalize_name,
    build_speaker_lookup,
)


def test_clean_str_handles_nan_and_none():
    assert clean_str(float('nan')) == ''
    assert clean_str(None) == ''
    assert clean_str('Hello') == 'Hello'


def test_extract_ace_director():
    assert extract_ace('Oracle ACE Director') == 'Oracle ACE Director'


def test_extract_ace_none():
    assert extract_ace('Pythian - Senior Database Consultant') == ''


def test_extract_ace_empty():
    assert extract_ace('') == ''


def test_extract_company_dash_format():
    assert extract_company('Pythian - Senior Database Consultant') == 'Pythian'


def test_extract_company_at_format():
    assert extract_company('Senior Database Consultant at Pythian') == 'Pythian'


def test_normalize_name_collapses_whitespace_and_case():
    assert normalize_name('  Hector Joaquin   Andrade Rodriguez ') == 'hector joaquin andrade rodriguez'


def test_build_speaker_lookup_joins_tagline_and_bio():
    accepted_rows = [
        {'FirstName': 'Eugenio', 'LastName': 'Galiano', 'SessionId': '1152551'},
    ]
    orig_rows = [
        {'SessionId': '1152551', 'TagLine': 'Oracle - Distinguished Product Manager', 'Bio': 'Database expert.'},
    ]
    lookup = build_speaker_lookup(accepted_rows, orig_rows)
    assert lookup['eugenio galiano'] == {
        'company': 'Oracle',
        'bio': 'Database expert.',
        'oracle_ace': '',
    }


def test_build_speaker_lookup_missing_orig_row_yields_empty_fields():
    accepted_rows = [{'FirstName': 'Nobody', 'LastName': 'Here', 'SessionId': '999'}]
    lookup = build_speaker_lookup(accepted_rows, [])
    assert lookup['nobody here'] == {'company': '', 'bio': '', 'oracle_ace': ''}


from openpyxl import Workbook
from openpyxl.styles import PatternFill
from export_agenda import is_keynote_row, extract_city_sessions


def _make_test_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Mexico'
    ws['A2'] = 'Horario'
    ws['B2'] = 'APEX'
    ws['C2'] = 'Base de Datos'
    ws['D2'] = 'Cloud & Diff. Topics'
    ws['E2'] = 'Inteligencia Artificial'
    ws.merge_cells('B3:E3')
    ws['B3'] = 'Keynote Title\nKeynote Speaker  (KEYNOTE)'
    ws['A4'] = '09:45 – 10:30'
    ws['B4'] = 'APEX Session\nSpeaker One'
    return ws


def test_is_keynote_row_true_for_merged_row():
    assert is_keynote_row(_make_test_sheet(), 3) is True


def test_is_keynote_row_false_for_normal_row():
    assert is_keynote_row(_make_test_sheet(), 4) is False


def test_extract_city_sessions_reads_keynote_and_track_session():
    ws = _make_test_sheet()
    green = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
    ws['B3'].fill = green
    ws['B4'].fill = green

    entries = extract_city_sessions(ws)

    assert entries == [
        {
            'city': 'Mexico', 'time_slot': None, 'track': None, 'is_keynote': True,
            'title': 'Keynote Title', 'speaker_name': 'Keynote Speaker', 'fill_rgb': 'FFC6EFCE',
        },
        {
            'city': 'Mexico', 'time_slot': '09:45 – 10:30', 'track': 'APEX', 'is_keynote': False,
            'title': 'APEX Session', 'speaker_name': 'Speaker One', 'fill_rgb': 'FFC6EFCE',
        },
    ]


def test_extract_city_sessions_skips_empty_cells():
    ws = _make_test_sheet()
    ws['B3'] = None
    entries = extract_city_sessions(ws)
    assert entries == [
        {
            'city': 'Mexico', 'time_slot': '09:45 – 10:30', 'track': 'APEX', 'is_keynote': False,
            'title': 'APEX Session', 'speaker_name': 'Speaker One', 'fill_rgb': '00000000',
        },
    ]


def _make_sheet_with_lunch_break():
    ws = _make_test_sheet()
    ws['A8'] = '  ALMUERZO  /  LUNCH BREAK     13:00 – 14:30'
    ws.merge_cells('A8:E8')
    return ws


def test_is_keynote_row_false_for_lunch_break_merge():
    ws = _make_sheet_with_lunch_break()
    assert is_keynote_row(ws, 8) is False


from export_agenda import build_public_sessions


def test_build_public_sessions_filters_unconfirmed_and_enriches():
    raw_entries = [
        {
            'city': 'Mexico', 'time_slot': '09:45 – 10:30', 'track': 'APEX', 'is_keynote': False,
            'title': 'APEX Session', 'speaker_name': 'Eugenio Galiano', 'fill_rgb': 'FFC6EFCE',
        },
        {
            'city': 'Mexico', 'time_slot': '10:30 – 11:15', 'track': 'APEX', 'is_keynote': False,
            'title': 'Unconfirmed Session', 'speaker_name': 'Someone Pending', 'fill_rgb': 'FFFFF2CC',
        },
    ]
    speaker_lookup = {
        'eugenio galiano': {'company': 'Oracle', 'bio': 'Bio text.', 'oracle_ace': ''},
    }

    result = build_public_sessions(raw_entries, speaker_lookup)

    assert result == [
        {
            'city': 'Mexico', 'time_slot': '09:45 – 10:30', 'track': 'APEX', 'is_keynote': False,
            'title': 'APEX Session', 'speaker_name': 'Eugenio Galiano',
            'speaker_company': 'Oracle', 'speaker_bio': 'Bio text.', 'oracle_ace': None,
        },
    ]


def test_build_public_sessions_unmatched_speaker_gets_empty_enrichment():
    raw_entries = [
        {
            'city': 'Mexico', 'time_slot': '09:45 – 10:30', 'track': 'APEX', 'is_keynote': False,
            'title': 'Mystery Session', 'speaker_name': 'Unknown Person', 'fill_rgb': 'FFC6EFCE',
        },
    ]
    result = build_public_sessions(raw_entries, {})
    assert result == [
        {
            'city': 'Mexico', 'time_slot': '09:45 – 10:30', 'track': 'APEX', 'is_keynote': False,
            'title': 'Mystery Session', 'speaker_name': 'Unknown Person',
            'speaker_company': '', 'speaker_bio': '', 'oracle_ace': None,
        },
    ]


from export_agenda import parse_custom_format_cell, extract_custom_format_sessions


def test_parse_custom_format_cell_speaker_then_title():
    text = 'Connor McDonald\nThe Humbling Experience of being an AI Newbie'
    assert parse_custom_format_cell(text) == {
        'title': 'The Humbling Experience of being an AI Newbie',
        'speaker_name': 'Connor McDonald',
        'is_keynote': False,
    }


def test_parse_custom_format_cell_keynote_prefix():
    text = 'Markus Michalewicz\nKeynote: Using Artificial Intelligence in Oracle AI Database'
    assert parse_custom_format_cell(text) == {
        'title': 'Using Artificial Intelligence in Oracle AI Database',
        'speaker_name': 'Markus Michalewicz',
        'is_keynote': True,
    }


def test_parse_custom_format_cell_single_line_returns_none():
    assert parse_custom_format_cell('Coffee Break') is None
    assert parse_custom_format_cell('Lunch Break') is None
    assert parse_custom_format_cell('Registro') is None
    assert parse_custom_format_cell('Apertura') is None


def _make_uruguay_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Uruguay'
    ws['A1'] = 'UYOUG LAOUC Tour 2026 | Lunes 31 Agosto'
    ws['A2'] = 'Horario'
    ws['B2'] = 'Sala D13'
    ws['C2'] = 'Auditorio'
    ws['A3'] = '08:30 - 09:00'
    ws['B3'] = 'Registro'
    ws['A4'] = '09:15 - 10:00'
    ws['B4'] = 'Markus Michalewicz\nKeynote: Using Artificial Intelligence in Oracle AI Database'
    ws['A5'] = '10:05 - 10:45'
    ws['B5'] = 'Connor McDonald\nThe Humbling Experience of being an AI Newbie'
    ws['C5'] = 'Richard Martens\nProperty Graphs in Oracle'
    return ws


def test_extract_custom_format_sessions_skips_logistics_and_flags_keynote():
    ws = _make_uruguay_sheet()
    entries = extract_custom_format_sessions(ws)
    assert entries == [
        {
            'city': 'Uruguay', 'time_slot': None, 'track': None, 'is_keynote': True,
            'title': 'Using Artificial Intelligence in Oracle AI Database',
            'speaker_name': 'Markus Michalewicz', 'fill_rgb': GREEN_FILL,
        },
        {
            'city': 'Uruguay', 'time_slot': '10:05 - 10:45', 'track': 'Sala D13', 'is_keynote': False,
            'title': 'The Humbling Experience of being an AI Newbie',
            'speaker_name': 'Connor McDonald', 'fill_rgb': GREEN_FILL,
        },
        {
            'city': 'Uruguay', 'time_slot': '10:05 - 10:45', 'track': 'Auditorio', 'is_keynote': False,
            'title': 'Property Graphs in Oracle',
            'speaker_name': 'Richard Martens', 'fill_rgb': GREEN_FILL,
        },
    ]


from export_agenda import parse_chile_regular_cell, parse_chile_keynote_cell, extract_chile_sessions


def test_parse_chile_regular_cell_splits_on_padding():
    text = 'Patching like a Pro - 2026 Edition                                                   Rodrigo Jorge'
    assert parse_chile_regular_cell(text) == {
        'title': 'Patching like a Pro - 2026 Edition',
        'speaker_name': 'Rodrigo Jorge',
    }


def test_parse_chile_regular_cell_logistics_returns_none():
    assert parse_chile_regular_cell('Coffee Break') is None
    assert parse_chile_regular_cell('Closing') is None


def test_parse_chile_keynote_cell():
    text = '▶  KEYNOTE  —  Engineering Resilience: The High Availability Mindset - Francisco Munoz Alvarez'
    assert parse_chile_keynote_cell(text) == {
        'title': 'Engineering Resilience: The High Availability Mindset',
        'speaker_name': 'Francisco Munoz Alvarez',
    }


def _make_chile_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Chile'
    ws['A2'] = 'Horario'
    ws['B2'] = 'Base de Datos'
    ws['C2'] = 'APEX / Desarrollo'
    ws.merge_cells('B3:E3')
    ws['B3'] = '▶  KEYNOTE  —  Engineering Resilience: The High Availability Mindset - Francisco Munoz Alvarez'
    ws['A4'] = '09:50 – 10:35'
    ws['B4'] = '100% Live Demo Is Back!                                            Connor McDonald'
    ws['C4'] = 'Búsqueda declarativa en APEX                                               Niall McPhillips'
    ws['A5'] = '10:35 - 11:05'
    ws.merge_cells('B5:E5')
    ws['B5'] = 'Coffee Break'
    return ws


def test_extract_chile_sessions_treats_every_cell_as_confirmed():
    entries = extract_chile_sessions(_make_chile_sheet())
    assert entries == [
        {
            'city': 'Chile', 'time_slot': None, 'track': None, 'is_keynote': True,
            'title': 'Engineering Resilience: The High Availability Mindset',
            'speaker_name': 'Francisco Munoz Alvarez', 'fill_rgb': GREEN_FILL,
        },
        {
            'city': 'Chile', 'time_slot': '09:50 – 10:35', 'track': 'Base de Datos', 'is_keynote': False,
            'title': '100% Live Demo Is Back!', 'speaker_name': 'Connor McDonald', 'fill_rgb': GREEN_FILL,
        },
        {
            'city': 'Chile', 'time_slot': '09:50 – 10:35', 'track': 'APEX / Desarrollo', 'is_keynote': False,
            'title': 'Búsqueda declarativa en APEX', 'speaker_name': 'Niall McPhillips', 'fill_rgb': GREEN_FILL,
        },
    ]


from export_agenda import parse_brazil_cell, extract_brazil_sessions


def test_parse_brazil_cell_regular_session():
    text = 'Jayson Hanes\nDistinguished Product Manager - Oracle\nOracle APEX Generative Development\nInglês'
    assert parse_brazil_cell(text) == {
        'title': 'Oracle APEX Generative Development',
        'speaker_name': 'Jayson Hanes',
        'is_keynote': False,
    }


def test_parse_brazil_cell_keynote_with_room():
    text = (
        'Key Note\nMike Dietrich\nVP Product Management and Development - Oracle\n'
        'Oracle 26ai: When the Database Becomes Intelligent and Autonomous\nEnglish\nAuditório'
    )
    assert parse_brazil_cell(text) == {
        'title': 'Oracle 26ai: When the Database Becomes Intelligent and Autonomous',
        'speaker_name': 'Mike Dietrich',
        'is_keynote': True,
    }


def test_parse_brazil_cell_co_presenters_extra_role_line():
    text = (
        'Mike Dietrich & Harsh Gupta\nVP Product Management - Oracle /\n'
        'VP Global Oracle Cloud and Database Services - Deutsche Bank\n'
        "Deutsche Bank's ExaC@C Journey: Transforming more than 8,000 databases to Oracle cloud\nInglês"
    )
    assert parse_brazil_cell(text) == {
        'title': "Deutsche Bank's ExaC@C Journey: Transforming more than 8,000 databases to Oracle cloud",
        'speaker_name': 'Mike Dietrich & Harsh Gupta',
        'is_keynote': False,
    }


def test_parse_brazil_cell_logistics_returns_none():
    assert parse_brazil_cell('Registro ') is None
    assert parse_brazil_cell('Lunch') is None
    assert parse_brazil_cell('Coffee Break') is None


def _make_brazil_sheet():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Brazil'
    ws['A3'] = 'Horário'
    ws['B3'] = 'APEX & Development'
    ws['C3'] = 'Database'
    ws['A4'] = 'Dia todo'
    ws['B4'] = 'Registro'
    ws['A7'] = '08:45-09:45AM'
    ws['B7'] = (
        'Key Note\nMike Dietrich\nVP Product Management - Oracle\n'
        'Oracle 26ai: When the Database Becomes Intelligent and Autonomous\nEnglish\nAuditório'
    )
    ws['A8'] = '10:00-10:45AM'
    ws['B8'] = 'Jayson Hanes\nDistinguished Product Manager - Oracle\nOracle APEX Generative Development\nInglês'
    return ws


def test_extract_brazil_sessions_skips_logistics_and_flags_keynote():
    ws = _make_brazil_sheet()
    entries = extract_brazil_sessions(ws)
    assert entries == [
        {
            'city': 'Brazil', 'time_slot': None, 'track': None, 'is_keynote': True,
            'title': 'Oracle 26ai: When the Database Becomes Intelligent and Autonomous',
            'speaker_name': 'Mike Dietrich', 'fill_rgb': GREEN_FILL,
        },
        {
            'city': 'Brazil', 'time_slot': '10:00-10:45AM', 'track': 'APEX & Development', 'is_keynote': False,
            'title': 'Oracle APEX Generative Development', 'speaker_name': 'Jayson Hanes', 'fill_rgb': GREEN_FILL,
        },
    ]
